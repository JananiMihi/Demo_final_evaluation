"""
build_table2.py
Builds Table 2 — PCA component loadings for LME-retained components.
Uses REAL computed values from pca_hr_loadings.csv and pca_bp_loadings.csv.
|loading| >= 0.30 threshold (matching paper convention).
"""
import os, warnings
warnings.filterwarnings("ignore")
import pandas as pd
import numpy as np

try:
    from openpyxl import Workbook
    from openpyxl.styles import (PatternFill, Font, Alignment, Border, Side,
                                  GradientFill)
    from openpyxl.utils import get_column_letter
    HAS_OPX = True
except ImportError:
    HAS_OPX = False
    print("openpyxl not found — saving as CSV instead")

WORK_DIR = os.path.dirname(os.path.abspath(__file__))
THRESH   = 0.30   # |loading| threshold

# ── 1. Load loading matrices ──────────────────────────────────────────────────
hr_load = pd.read_csv(os.path.join(WORK_DIR, "pca_hr_loadings.csv"), index_col=0)
bp_load = pd.read_csv(os.path.join(WORK_DIR, "pca_bp_loadings.csv"), index_col=0)
var_df  = pd.read_csv(os.path.join(WORK_DIR, "pca_variance_explained.csv"))

hr_var = dict(zip(var_df[var_df["Domain"]=="HR"]["PC"],
                  var_df[var_df["Domain"]=="HR"]["VarExp_pct"]))
bp_var = dict(zip(var_df[var_df["Domain"]=="BP"]["PC"],
                  var_df[var_df["Domain"]=="BP"]["VarExp_pct"]))

# ── 2. LME-retained PCs (from parsimonious model fixed effects) ──────────────
# Confirmed from lme_fixed_effects_parsimonious.csv
LME_HR_PCS = ["HR_PC2", "HR_PC3", "HR_PC5", "HR_PC9", "HR_PC10"]
LME_BP_PCS = ["BP_PC3", "BP_PC4", "BP_PC5", "BP_PC6"]

# Feature name mapping (raw → clean label for table)
FEAT_MAP = {
    # HR
    "EX_RMSSD_ms":             "EX_RMSSD_ms",
    "EX_SDNN_ms_log":          "EX_SDNN_ms (log)",
    "EX_pNNxx_pct":            "EX_pNNxx_pct",
    "EX_Mean_RR_ms":           "EX_Mean_RR_ms",
    "EX_Mean_HR_bpm":          "EX_Mean_HR_bpm",
    "EX_RR_tri_index":         "EX_RR_tri_index",
    "EX_TINN_ms_log":          "EX_TINN_ms (log)",
    "EX_LF_abs_FFT_log":       "EX_LF_abs_FFT (log)",
    "EX_HF_abs_FFT_log":       "EX_HF_abs_FFT (log)",
    "EX_LF_HF_ratio_FFT":      "EX_LF_HF_ratio_FFT",
    "EX_VLF_abs_FFT_log":      "EX_VLF_abs_FFT (log)",
    "EX_Total_power_FFT_log":  "EX_Total_power_FFT (log)",
    "EX_LF_nu_FFT":            "EX_LF_nu_FFT",
    "EX_HF_nu_FFT":            "EX_HF_nu_FFT",
    "EX_SD1_ms_log":           "EX_SD1_ms (log)",
    "EX_SD2_ms_log":           "EX_SD2_ms (log)",
    "EX_SD2_SD1_ratio_log":    "EX_SD2_SD1_ratio (log)",
    "EX_SampEn_log":           "EX_SampEn (log)",
    "EX_DFA_alpha1":           "EX_DFA_alpha1",
    "EX_RQA_RecurrenceRate":   "EX_RQA_RecurrenceRate",
    "EX_RQA_Determinism":      "EX_RQA_Determinism",
    "REST_RMSSD_ms_log":       "REST_RMSSD_ms (log)",
    "REST_SDNN_ms_log":        "REST_SDNN_ms (log)",
    "REST_pNNxx_pct":          "REST_pNNxx_pct",
    "REST_Mean_RR_ms":         "REST_Mean_RR_ms",
    "REST_Mean_HR_bpm":        "REST_Mean_HR_bpm",
    "REST_LF_abs_FFT_log":     "REST_LF_abs_FFT (log)",
    "REST_HF_abs_FFT_log":     "REST_HF_abs_FFT (log)",
    "REST_LF_HF_ratio_FFT":    "REST_LF_HF_ratio_FFT",
    "REST_Total_power_FFT_log":"REST_Total_power_FFT (log)",
    "REST_LF_nu_FFT":          "REST_LF_nu_FFT",
    "REST_HF_nu_FFT":          "REST_HF_nu_FFT",
    "REST_SD1_ms_log":         "REST_SD1_ms (log)",
    "REST_SD2_ms_log":         "REST_SD2_ms (log)",
    "REST_SampEn":             "REST_SampEn",
    "REST_DFA_alpha1":         "REST_DFA_alpha1",
    "nabla_h_ex":              "nabla_h_ex",
    "nabla_h_rest":            "nabla_h_rest",
    "GLOBAL_EX_RMSSD_ms_log":  "GLOBAL_EX_RMSSD_ms (log)",
    "GLOBAL_EX_SDNN_ms_log":   "GLOBAL_EX_SDNN_ms (log)",
    "GLOBAL_REST_RMSSD_ms_log":"GLOBAL_REST_RMSSD_ms (log)",
    "GLOBAL_REST_SDNN_ms_log": "GLOBAL_REST_SDNN_ms (log)",
    "GLOBAL_nabla_h_ex":       "GLOBAL_nabla_h_ex",
    "GLOBAL_nabla_h_rest":     "GLOBAL_nabla_h_rest",
    # BP
    "SBP_rest_pre_mmHg":   "SBP_rest_pre_mmHg",
    "DBP_rest_pre_mmHg":   "DBP_rest_pre_mmHg",
    "MAP_rest_pre_mmHg":   "MAP_rest_pre_mmHg",
    "PP_rest_pre_mmHg":    "PP_rest_pre_mmHg",
    "SBP_post_set_mmHg":   "SBP_post_set_mmHg",
    "DBP_post_set_mmHg":   "DBP_post_set_mmHg",
    "MAP_post_set_mmHg":   "MAP_post_set_mmHg",
    "delta_SBP_mmHg":      "delta_SBP_mmHg",
    "SBP_mean_e":          "SBP_mean_e",
    "DBP_mean_e":          "DBP_mean_e",
    "MAP_e":               "MAP_e",
    "Delta_SBP_e":         "Delta_SBP_e",
    "Delta_DBP_e":         "Delta_DBP_e",
    "ARV_SBP_mmHg":        "ARV_SBP_mmHg",
    "SV_SBP_mmHg":         "SV_SBP_mmHg",
    "ARV_DBP_mmHg":        "ARV_DBP_mmHg",
    "SV_DBP_mmHg":         "SV_DBP_mmHg",
    "VIM_SBP_mmHg":        "VIM_SBP_mmHg",
    "VIM_DBP_mmHg":        "VIM_DBP_mmHg",
}

# ── 3. Extract Table 2 rows ───────────────────────────────────────────────────
def get_high_loadings(load_df, pc_name, thresh=0.30):
    if pc_name not in load_df.columns:
        return []
    col = load_df[pc_name]
    high = col[col.abs() >= thresh].sort_values(key=abs, ascending=False)
    return [(FEAT_MAP.get(feat, feat), round(float(v), 4)) for feat, v in high.items()]

table2_rows = []

for pc in LME_HR_PCS:
    items = get_high_loadings(hr_load, pc)
    var_pct = hr_var.get(pc, None)
    pc_label = f"{pc} ({var_pct:.2f}% var)" if var_pct else pc
    if items:
        for i, (feat, val) in enumerate(items):
            table2_rows.append({
                "Component":      pc if i == 0 else "",
                "Var Explained %": round(var_pct, 2) if (i == 0 and var_pct) else "",
                "Raw Feature":    feat,
                "Loading":        val,
                "|Loading|":      abs(val),
            })
    else:
        table2_rows.append({
            "Component":      pc,
            "Var Explained %": round(var_pct, 2) if var_pct else "",
            "Raw Feature":    "(no feature with |loading| ≥ 0.30)",
            "Loading":        None,
            "|Loading|":      None,
        })
    table2_rows.append({"Component":"","Var Explained %":"","Raw Feature":"","Loading":"","|Loading|":""})

for pc in LME_BP_PCS:
    items = get_high_loadings(bp_load, pc)
    var_pct = bp_var.get(pc, None)
    if items:
        for i, (feat, val) in enumerate(items):
            table2_rows.append({
                "Component":      pc if i == 0 else "",
                "Var Explained %": round(var_pct, 2) if (i == 0 and var_pct) else "",
                "Raw Feature":    feat,
                "Loading":        val,
                "|Loading|":      abs(val),
            })
    else:
        table2_rows.append({
            "Component":      pc,
            "Var Explained %": round(var_pct, 2) if var_pct else "",
            "Raw Feature":    "(no feature with |loading| ≥ 0.30)",
            "Loading":        None,
            "|Loading|":      None,
        })
    table2_rows.append({"Component":"","Var Explained %":"","Raw Feature":"","Loading":"","|Loading|":""})

df_t2 = pd.DataFrame(table2_rows)

# ── 4. Print console preview ──────────────────────────────────────────────────
print("\n" + "="*70)
print("TABLE 2 — PCA Component Loadings for LME-Retained Components")
print(f"(Only |loading| ≥ {THRESH} shown)")
print("="*70)
cur = None
for _, r in df_t2.iterrows():
    if r["Component"] and r["Component"] != cur:
        cur = r["Component"]
        print(f"\n  {r['Component']}  (Var explained: {r['Var Explained %']}%)")
        print(f"  {'Raw Feature':<40s}  {'Loading':>8s}  {'|Loading|':>9s}")
        print("  " + "-"*60)
    if r["Raw Feature"]:
        sign = "+" if isinstance(r["Loading"], float) and r["Loading"] > 0 else ""
        load_str = f"{sign}{r['Loading']:.4f}" if isinstance(r["Loading"], float) else str(r["Loading"])
        abs_str  = f"{r['|Loading|']:.4f}" if isinstance(r["|Loading|"], float) else ""
        print(f"  {r['Raw Feature']:<40s}  {load_str:>8s}  {abs_str:>9s}")

# ── 5. Build Excel workbook ───────────────────────────────────────────────────
if HAS_OPX:
    wb = Workbook()
    ws = wb.active
    ws.title = "Table 2 - PCA Loadings"

    # ── Color palette ──────────────────────────────────────────────────────────
    C_HEADER   = "1E3A5F"   # dark navy
    C_HR_GRP   = "0F2D4A"   # deep blue-grey for HR rows
    C_BP_GRP   = "2D1B4A"   # deep purple for BP rows
    C_SPACER   = "0F172A"   # near-black for spacer rows
    C_H_TXT    = "FFFFFF"
    C_POS_HIGH = "1A4731"   # dark green bg — strong positive (>0.40)
    C_POS_MED  = "1C3A20"   # medium green
    C_NEG_HIGH = "3B1010"   # dark red bg — strong negative (<-0.40)
    C_NEG_MED  = "3A1C1C"   # medium red
    C_ZERO     = "1E293B"   # near-black for near-zero values

    def hdr_fill(hex_): return PatternFill("solid", fgColor=hex_)
    def thin_border():
        s = Side(border_style="thin", color="334155")
        return Border(left=s, right=s, top=s, bottom=s)
    def bold(size=10): return Font(bold=True, size=size, color="FFFFFF", name="Calibri")
    def reg(color="CBD5E1", size=9.5): return Font(size=size, color=color, name="Calibri")
    def num_font(v):
        if v is None or v == "": return reg()
        c = "4ADE80" if v > 0 else ("F87171" if v < 0 else "94A3B8")
        return Font(size=9.5, color=c, name="Courier New", bold=abs(v) >= 0.40)

    # ── Title / caption rows ───────────────────────────────────────────────────
    ws.merge_cells("A1:E1")
    ws["A1"] = "Table 2: PCA Component Loadings for LME-Retained HR and BP Components"
    ws["A1"].font = Font(bold=True, size=13, color="F8FAFC", name="Calibri")
    ws["A1"].alignment = Alignment(horizontal="center", vertical="center")
    ws["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws.row_dimensions[1].height = 22

    ws.merge_cells("A2:E2")
    ws["A2"] = f"Only loadings with |loading| ≥ {THRESH} are shown. Real computed values (N=234 rows, 63 participants). Full loading matrix available in supplementary material."
    ws["A2"].font = Font(italic=True, size=8.5, color="94A3B8", name="Calibri")
    ws["A2"].alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws["A2"].fill = PatternFill("solid", fgColor="0F172A")
    ws.row_dimensions[2].height = 28

    # ── Column headers ─────────────────────────────────────────────────────────
    headers = ["Component", "Var Exp (%)", "Raw Feature", "Loading", "|Loading|"]
    for col, h in enumerate(headers, 1):
        c = ws.cell(row=3, column=col, value=h)
        c.font = bold(10)
        c.fill = hdr_fill(C_HEADER)
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = thin_border()
    ws.row_dimensions[3].height = 18

    # ── Column widths ──────────────────────────────────────────────────────────
    ws.column_dimensions["A"].width = 22
    ws.column_dimensions["B"].width = 14
    ws.column_dimensions["C"].width = 38
    ws.column_dimensions["D"].width = 13
    ws.column_dimensions["E"].width = 13

    # ── Data rows ──────────────────────────────────────────────────────────────
    row_idx = 4
    for _, r in df_t2.iterrows():
        comp   = r["Component"]
        feat   = r["Raw Feature"]
        var_e  = r["Var Explained %"]
        loading= r["Loading"]
        abs_l  = r["|Loading|"]

        # Spacer row
        if not comp and not feat:
            for c in range(1, 6):
                cell = ws.cell(row=row_idx, column=c, value="")
                cell.fill = PatternFill("solid", fgColor=C_SPACER)
            ws.row_dimensions[row_idx].height = 5
            row_idx += 1
            continue

        is_hr = any(comp.startswith("HR") or r["Component"] == "" for _ in [1])
        # Determine group color
        cur_comp = comp if comp else ""
        is_hr_row = cur_comp.startswith("HR") if cur_comp else (
            row_idx > 0  # default to last group color
        )

        # Determine bg based on loading strength
        if isinstance(loading, float):
            if loading >= 0.40:    bg = C_POS_HIGH
            elif loading >= 0.30:  bg = C_POS_MED
            elif loading <= -0.40: bg = C_NEG_HIGH
            elif loading <= -0.30: bg = C_NEG_MED
            else:                  bg = C_ZERO
        else:
            bg = "161F2E"

        # Write cells
        cells_data = [
            (1, comp, "left", Font(bold=bool(comp), size=9.5, color="38BDF8" if comp.startswith("HR") else "A78BFA", name="Calibri") if comp else reg()),
            (2, f"{var_e}%" if var_e else "", "center", reg("64748B")),
            (3, feat, "left", reg("E2E8F0") if feat else reg("475569")),
            (4, f"{'+' if isinstance(loading,float) and loading>0 else ''}{loading:.4f}" if isinstance(loading,float) else "", "center", num_font(loading) if isinstance(loading,float) else reg()),
            (5, f"{abs_l:.4f}" if isinstance(abs_l,float) else "", "center", Font(size=9.5, color="F1F5F9", name="Courier New", bold=isinstance(abs_l,float) and abs_l>=0.40)),
        ]
        for col, val, align, font in cells_data:
            cell = ws.cell(row=row_idx, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=bg)
            cell.font  = font
            cell.alignment = Alignment(horizontal=align, vertical="center")
            cell.border = thin_border()
        ws.row_dimensions[row_idx].height = 16
        row_idx += 1

    # ── Legend ─────────────────────────────────────────────────────────────────
    row_idx += 1
    ws.merge_cells(f"A{row_idx}:E{row_idx}")
    ws.cell(row=row_idx, column=1).value = "COLOR KEY:  Dark green = strong positive loading (≥ 0.40)  |  Medium green = 0.30–0.39  |  Dark red = strong negative (≤ −0.40)  |  Medium red = −0.30 to −0.39"
    ws.cell(row=row_idx, column=1).font  = Font(italic=True, size=8, color="64748B", name="Calibri")
    ws.cell(row=row_idx, column=1).fill  = PatternFill("solid", fgColor="0F172A")
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 14

    row_idx += 1
    ws.merge_cells(f"A{row_idx}:E{row_idx}")
    ws.cell(row=row_idx, column=1).value = "LME-retained PCs: HR_PC2, HR_PC3, HR_PC5, HR_PC9, HR_PC10 (HR domain)  |  BP_PC3, BP_PC4, BP_PC5, BP_PC6 (BP domain)"
    ws.cell(row=row_idx, column=1).font  = Font(italic=True, size=8, color="64748B", name="Calibri")
    ws.cell(row=row_idx, column=1).fill  = PatternFill("solid", fgColor="0F172A")
    ws.cell(row=row_idx, column=1).alignment = Alignment(horizontal="left", vertical="center")
    ws.row_dimensions[row_idx].height = 14

    # ── Also add a "Full Loading Matrix" sheet for all 17 HR + 8 BP PCs ───────
    ws2 = wb.create_sheet("HR Full Loadings")
    ws2["A1"] = "HR Domain — Full Loading Matrix (44 features × 17 PCs)"
    ws2["A1"].font = Font(bold=True, size=11, color="38BDF8", name="Calibri")
    ws2["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws2.merge_cells("A1:S1")
    hr_cols = list(hr_load.columns)
    for ci, col in enumerate(["Feature"] + hr_cols, 1):
        c = ws2.cell(row=2, column=ci, value=col)
        c.font  = Font(bold=True, size=8.5, color="FFFFFF", name="Calibri")
        c.fill  = hdr_fill(C_HEADER)
        c.alignment = Alignment(horizontal="center")
    ws2.column_dimensions["A"].width = 35
    for ci in range(2, len(hr_cols)+2):
        ws2.column_dimensions[get_column_letter(ci)].width = 9
    for ri, (feat, row_s) in enumerate(hr_load.iterrows(), 3):
        ws2.cell(row=ri, column=1, value=FEAT_MAP.get(feat,feat)).font = reg("94A3B8", 8.5)
        ws2.cell(row=ri, column=1).fill = PatternFill("solid", fgColor="0F172A")
        for ci, v in enumerate(row_s.values, 2):
            v = float(v)
            if abs(v) >= 0.40:    bg2 = C_POS_HIGH if v>0 else C_NEG_HIGH
            elif abs(v) >= 0.25:  bg2 = C_POS_MED  if v>0 else C_NEG_MED
            else:                 bg2 = "0F172A"
            cell = ws2.cell(row=ri, column=ci,
                            value=round(v,4) if abs(v)>=0.05 else "")
            cell.font  = Font(size=8, color=("4ADE80" if v>0 else "F87171") if abs(v)>=0.10 else "334155",
                              name="Courier New")
            cell.fill  = PatternFill("solid", fgColor=bg2)
            cell.alignment = Alignment(horizontal="center")
        ws2.row_dimensions[ri].height = 14

    ws3 = wb.create_sheet("BP Full Loadings")
    ws3["A1"] = "BP Domain — Full Loading Matrix (19 features × 8 PCs)"
    ws3["A1"].font = Font(bold=True, size=11, color="A78BFA", name="Calibri")
    ws3["A1"].fill = PatternFill("solid", fgColor="0F172A")
    ws3.merge_cells("A1:J1")
    bp_cols = list(bp_load.columns)
    for ci, col in enumerate(["Feature"] + bp_cols, 1):
        c = ws3.cell(row=2, column=ci, value=col)
        c.font  = Font(bold=True, size=9, color="FFFFFF", name="Calibri")
        c.fill  = hdr_fill("4C1D95")
        c.alignment = Alignment(horizontal="center")
    ws3.column_dimensions["A"].width = 28
    for ci in range(2, len(bp_cols)+2):
        ws3.column_dimensions[get_column_letter(ci)].width = 10
    for ri, (feat, row_s) in enumerate(bp_load.iterrows(), 3):
        ws3.cell(row=ri, column=1, value=FEAT_MAP.get(feat,feat)).font = reg("94A3B8", 9)
        ws3.cell(row=ri, column=1).fill = PatternFill("solid", fgColor="0F172A")
        for ci, v in enumerate(row_s.values, 2):
            v = float(v)
            if abs(v) >= 0.40:    bg2 = C_POS_HIGH if v>0 else C_NEG_HIGH
            elif abs(v) >= 0.25:  bg2 = C_POS_MED  if v>0 else C_NEG_MED
            else:                 bg2 = "0F172A"
            cell = ws3.cell(row=ri, column=ci,
                            value=round(v,4) if abs(v)>=0.05 else "")
            cell.font  = Font(size=9, color=("4ADE80" if v>0 else "F87171") if abs(v)>=0.10 else "334155",
                              name="Courier New")
            cell.fill  = PatternFill("solid", fgColor=bg2)
            cell.alignment = Alignment(horizontal="center")
        ws3.row_dimensions[ri].height = 15

    # Save
    out_path = os.path.join(WORK_DIR, "Table2_PCA_Loadings_Real.xlsx")
    wb.save(out_path)
    print(f"\n✅ Saved → {out_path}")
    print(f"   Sheets: '{ws.title}', '{ws2.title}', '{ws3.title}'")

else:
    # CSV fallback
    out_path = os.path.join(WORK_DIR, "Table2_PCA_Loadings_Real.csv")
    df_t2.to_csv(out_path, index=False)
    print(f"\nSaved CSV → {out_path}")

print("\nDONE")
