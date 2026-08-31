"""
HRVBP-HyperNet Demographic Dashboard
=====================================
A desktop app: drag & drop the questionnaire .xlsx onto the window (or use
Browse) and it computes/displays, per participant:

    Participant ID, Weight, Height, BMI, Age,
    PSS_total, PSS_category, Stress_index,
    PSQI_global, PSQI_category, Sleep_index,
    Nutrition_index, Nutrition_items_answered,
    Wellness_score

None of these derived columns exist ready-made in the source sheet (the
sheet's own "PSS Index" / "Sleep index" / "Score" columns are broken —
inconsistent formula ranges, mostly blank). This app recomputes everything
from the raw item-level answers using standard scoring rules:

  BMI              = weight_kg / (height_m ** 2)

  PSS_total        = standard PSS-10 score (0-40).
                      Items already stored 0-4. Q4,5,7,8 are reverse-scored
                      (score = 4 - raw) before summing all 10 items.
  PSS_category      Low 0-13 / Moderate 14-26 / High 27-40
  Stress_index      PSS_total / 40 * 100  (0-100, higher = more stress)

  PSQI_global       standard 7-component PSQI (0-21):
                      C1 quality, C2 latency, C3 duration, C4 efficiency,
                      C5 disturbances, C6 medication use, C7 daytime
                      dysfunction. Each component 0-3.
  PSQI_category     Good (<=5) / Poor (>5)
  Sleep_index       (1 - PSQI_global / 21) * 100  (0-100, higher = better sleep)

  Nutrition_index   mean weekly-frequency score across all answered FFQ food
                      items (frequency wording mapped to approx times/week,
                      e.g. "Once a week"=1, "2-4 per week"=3, "Once a day"=7,
                      "6+ per day"=42 ...).
  Nutrition_items_answered  count of FFQ food items the participant answered
                      (out of the ~130 item food-frequency list).

  Wellness_score    average of three 0-100 sub-scores:
                      (100 - Stress_index), Sleep_index,
                      min(Nutrition_index / 20 * 100, 100)
                      -> single 0-100 composite, higher = better overall.

These are standard/typical scoring conventions, not the study's own
(missing) formulas — treat them as a reasonable default and adjust the
mapping tables below if your protocol defines them differently.

Requirements:
    pip install openpyxl tkinterdnd2
    (tkinterdnd2 is optional — without it, use the "Browse..." button
     instead of drag & drop)
"""

import os
import re
import sys
import csv
import statistics
import tkinter as tk
from tkinter import ttk, filedialog, messagebox

try:
    from tkinterdnd2 import DND_FILES, TkinterDnD
    HAS_DND = True
except ImportError:
    HAS_DND = False

import openpyxl

SHEET_NAME = "Form responses 1"

# ---------------------------------------------------------------------
# Column headers we need, exactly as they appear in row 1 of the sheet.
# Using header text (not fixed indices) makes this robust to column
# reordering.
# ---------------------------------------------------------------------
COL = {
    "id": "Participant ID",
    "age": "Age (years)",
    "gender": "Gender\n",
    "height": "Height (cm)",
    "weight": "Weight (kg)",
    "pss": [
        "PSS Q1. How often were you upset by something unexpected? ",
        "PSS Q2. How often did you feel you couldn't control important things in your life??",
        "PSS Q3. In the last month, how often have you felt nervous and 'stressed'?",
        "PSS Q4 \u2605. How often did you feel confident in handling your problems? ",
        "PSS Q5 \u2605. How often did you feel things were going well for you?  ",
        "PSS Q6.How often did you feel you couldn't handle everything you had to do?  ",
        "PSS Q7 \u2605. How often could you manage the annoying things in your life? ",
        "PSS Q8 \u2605. How often did you feel you were on top of things? ",
        "PSS Q9. How often were you angry about things you couldn't control?  ",
        "PSS Q10. In the last month, how often have you felt difficulties were piling up so high that you could not overcome them?",
    ],
    "pss_reverse": {3, 4, 6, 7},  # 0-based indices within the pss list above (Q4,5,7,8)
    "psqi_bed": "PSQI Q1. During the past month, what time have you usually gone to bed at night?",
    "psqi_latency_min": "PSQI Q2. During the past month, how long (in minutes) has it usually taken you to fall asleep each night?",
    "psqi_wake": "PSQI Q3. During the past month, what time have you usually gotten up in the morning?",
    "psqi_hours_sleep": "PSQI Q4. During the past month, how many hours of actual sleep did you get at night?",
    "psqi_5a": "PSQI Q5a) Cannot get to sleep within 30 minutes",
    "psqi_5_rest": [
        "PSQI Q5b) Wake up in the middle of the night or early morning",
        "PSQI Q5c) Have to get up to use the bathroom",
        "PSQI Q5d) Cannot breathe comfortably",
        "PSQI Q5e) Cough or snore loudly",
        "PSQI Q5f) Feel too cold",
        "PSQI Q5g) Feel too hot",
        "PSQI Q5h) Had bad dreams",
        "PSQI Q5i) Have pain",
        "PSQI Q5j) If other reason \u2014 how often during the past month have you had trouble sleeping because of this?",
    ],
    "psqi_quality": "PSQI Q6. During the past month, how would you rate your sleep quality overall?",
    "psqi_meds": "PSQI Q7. During the past month, how often have you taken medicine to help you sleep (prescribed or \"over the counter\")?",
    "psqi_daytime_trouble": "PSQI Q8. During the past month, how often have you had trouble staying awake while driving, eating meals, or engaging in social activity?",
    "psqi_daytime_energy": "PSQI Q9. How often did you feel low energy or motivation to do your daily tasks because of poor sleep?",
}

# The block of individual FFQ food-frequency items (from "Beef" to
# "Tofu, soya meat, TVP, Vegeburger") used for the nutrition index.
NUTRITION_FIRST_HEADER = "Beef"
NUTRITION_LAST_HEADER = "Tofu, soya meat, TVP, Vegeburger"

# ---------------------------------------------------------------------
# Frequency word -> numeric mappings
# ---------------------------------------------------------------------
FREQ_0_3 = {
    "not during the past month": 0,
    "less than once a week": 1,
    "once or twice a week": 2,
    "three or more times a week": 3,
}

QUALITY_0_3 = {
    "very good": 0,
    "fairly good": 1,
    "fairly bad": 2,
    "very bad": 3,
}

PROBLEM_0_3 = {
    "no problem at all": 0,
    "only a very slight problem": 1,
    "somewhat of a problem": 2,
    "a very big problem": 3,
}

FFQ_TIMES_PER_WEEK = {
    "never or less than once a month": 0,
    "1\u20133 per month": 0.5,
    "once a week": 1,
    "2\u20134 per week": 3,
    "5\u20136 per week": 5.5,
    "once a day": 7,
    "2\u20133 per day": 17.5,
    "4\u20135 per day": 31.5,
    "6+ per day": 42,
}


def norm(s):
    return str(s).strip().lower() if s is not None else ""


def to_number(v):
    """Coerce a cell value to a float, tolerating stray text like '05',
    ' 30 ', '30 minutes', or '7-8'. Returns None if no number is found."""
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return float(v)
    s = str(v).strip()
    if not s:
        return None
    m = re.search(r'-?\d+(?:\.\d+)?', s)
    if not m:
        return None
    try:
        return float(m.group())
    except ValueError:
        return None


def score_q2_minutes(minutes):
    minutes = to_number(minutes)
    if minutes is None:
        return None
    if minutes <= 15:
        return 0
    if minutes <= 30:
        return 1
    if minutes <= 60:
        return 2
    return 3


def score_hours(hours):
    hours = to_number(hours)
    if hours is None:
        return None
    if hours > 7:
        return 0
    if hours >= 6:
        return 1
    if hours >= 5:
        return 2
    return 3


def collapse_0_6_to_0_3(total):
    if total is None:
        return None
    if total == 0:
        return 0
    if total <= 2:
        return 1
    if total <= 4:
        return 2
    return 3


def collapse_disturbance(total):
    if total is None:
        return None
    if total == 0:
        return 0
    if total <= 9:
        return 1
    if total <= 18:
        return 2
    return 3


def parse_clock(value):
    """Best-effort parse of a free-text/numeric bedtime or wake-time entry
    into (decimal_hour, am_pm) where am_pm is 'A'/'P'/None if not stated.
    Source data mixes formats like '11.30PM', '11.00 PM', '1:00 AM',
    '7 00 AM', and bare decimal hours like 8.3 (meaning ~8:18)."""
    if value is None or (isinstance(value, str) and not value.strip()):
        return None, None
    if isinstance(value, (int, float)):
        hour = int(value)
        minutes = (value - hour) * 60
        return hour + minutes / 60, None
    s = str(value).strip()
    m = re.search(r'(\d{1,2})[.:\s]+(\d{2})\s*([APap])', s)
    if m:
        hour, minute, ampm = int(m.group(1)), int(m.group(2)), m.group(3).upper()
    else:
        m = re.search(r'(\d{1,2})\s*([APap])', s)
        if not m:
            return None, None
        hour, minute, ampm = int(m.group(1)), 0, m.group(2).upper()
    if ampm == "P" and hour != 12:
        hour += 12
    if ampm == "A" and hour == 12:
        hour = 0
    return hour + minute / 60, ampm


def resolve_bed_hour(raw):
    hour, ampm = parse_clock(raw)
    if hour is None:
        return None
    if ampm is None:
        # No AM/PM stated (bare decimal entry): bedtimes 8-11 -> PM,
        # bedtimes 1-7 -> treated as already past midnight (AM).
        h = int(hour)
        if 8 <= h <= 11:
            hour += 12
    return hour % 24


def resolve_wake_hour(raw):
    hour, ampm = parse_clock(raw)
    if hour is None:
        return None
    if ampm is None:
        # No AM/PM stated: wake times are almost always AM; a bare "12"
        # is treated as noon (12 PM) since that's an unusually late wake.
        h = int(hour)
        if h == 12:
            hour += 12
    return hour % 24


def efficiency_component(bed_raw, wake_raw, hours_slept):
    """Estimate habitual sleep efficiency component (0-3) from bed/wake
    clock-time entries and reported hours of actual sleep. Best-effort:
    source bed/wake times are freely typed and inconsistently formatted."""
    bed = resolve_bed_hour(bed_raw)
    wake = resolve_wake_hour(wake_raw)
    hours_slept = to_number(hours_slept)
    if bed is None or wake is None or hours_slept is None:
        return None
    time_in_bed = (wake - bed) % 24
    if time_in_bed <= 0:
        return None
    efficiency = (hours_slept / time_in_bed) * 100
    if efficiency >= 85:
        return 0
    if efficiency >= 75:
        return 1
    if efficiency >= 65:
        return 2
    return 3


def compute_pss(row, headers_idx):
    values = []
    for i, q in enumerate(COL["pss"]):
        idx = headers_idx.get(q)
        v = to_number(row[idx]) if idx is not None else None
        if v is None:
            continue
        if i in COL["pss_reverse"]:
            v = 4 - v
        values.append(v)
    if len(values) < 8:  # allow up to 2 missing answers, prorated
        return None
    return round(sum(values) * 10 / len(values))


def pss_category(total):
    if total is None:
        return None
    if total <= 13:
        return "Low"
    if total <= 26:
        return "Moderate"
    return "High"


def compute_psqi(row, headers_idx):
    def get(key):
        idx = headers_idx.get(COL[key])
        return row[idx] if idx is not None else None

    # Component 1: subjective quality
    c1 = QUALITY_0_3.get(norm(get("psqi_quality")))

    # Component 2: latency
    q2_score = score_q2_minutes(get("psqi_latency_min"))
    q5a_score = FREQ_0_3.get(norm(get("psqi_5a")))
    if q2_score is not None and q5a_score is not None:
        c2 = collapse_0_6_to_0_3(q2_score + q5a_score)
    else:
        c2 = None

    # Component 3: duration
    c3 = score_hours(get("psqi_hours_sleep"))

    # Component 4: habitual sleep efficiency
    c4 = efficiency_component(get("psqi_bed"), get("psqi_wake"), get("psqi_hours_sleep"))

    # Component 5: disturbances (5b through 5j, 9 sub-items)
    disturb_scores = []
    for q in COL["psqi_5_rest"]:
        idx = headers_idx.get(q)
        if idx is None:
            continue
        v = FREQ_0_3.get(norm(row[idx]))
        if v is not None:
            disturb_scores.append(v)
    c5 = collapse_disturbance(sum(disturb_scores)) if disturb_scores else None

    # Component 6: sleep medication use
    c6 = FREQ_0_3.get(norm(get("psqi_meds")))

    # Component 7: daytime dysfunction
    trouble = FREQ_0_3.get(norm(get("psqi_daytime_trouble")))
    energy = PROBLEM_0_3.get(norm(get("psqi_daytime_energy")))
    if trouble is not None and energy is not None:
        c7 = collapse_0_6_to_0_3(trouble + energy)
    else:
        c7 = None

    components = [c1, c2, c3, c4, c5, c6, c7]
    known = [c for c in components if c is not None]
    # Require at least 5 of 7 components (missing ones scored as 0) so a
    # single unparseable field (usually the free-typed bed/wake time)
    # doesn't blank out the whole score.
    if len(known) < 5:
        return None
    return sum(c if c is not None else 0 for c in components)


def psqi_category(total):
    if total is None:
        return None
    return "Good" if total <= 5 else "Poor"


def compute_nutrition(row, headers, headers_idx):
    start = headers.index(NUTRITION_FIRST_HEADER)
    end = headers.index(NUTRITION_LAST_HEADER)
    freqs = []
    answered = 0
    for idx in range(start, end + 1):
        v = row[idx]
        if v is None or str(v).strip() == "":
            continue
        f = FFQ_TIMES_PER_WEEK.get(norm(v))
        if f is not None:
            freqs.append(f)
            answered += 1
    nutrition_index = round(statistics.mean(freqs), 2) if freqs else None
    return nutrition_index, answered


def wellness_score(stress_index, sleep_index, nutrition_index):
    parts = []
    if stress_index is not None:
        parts.append(100 - stress_index)
    if sleep_index is not None:
        parts.append(sleep_index)
    if nutrition_index is not None:
        parts.append(min(nutrition_index / 20 * 100, 100))
    if not parts:
        return None
    return round(statistics.mean(parts), 1)


def clean_id(raw_id):
    if raw_id is None:
        return ""
    return str(raw_id).split("(")[0].strip()


def process_workbook(path):
    wb = openpyxl.load_workbook(path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Sheet '{SHEET_NAME}' not found. Sheets: {wb.sheetnames}")
    ws = wb[SHEET_NAME]

    headers = [c.value for c in next(ws.iter_rows(min_row=1, max_row=1))]
    headers_idx = {h: i for i, h in enumerate(headers) if h is not None}

    results = []
    for row_cells in ws.iter_rows(min_row=2, values_only=True):
        row = list(row_cells)
        pid = clean_id(row[headers_idx.get(COL["id"], 0)])
        if not pid:
            continue

        age = to_number(row[headers_idx.get(COL["age"])]) if COL["age"] in headers_idx else None
        height = to_number(row[headers_idx.get(COL["height"])]) if COL["height"] in headers_idx else None
        weight = to_number(row[headers_idx.get(COL["weight"])]) if COL["weight"] in headers_idx else None

        bmi = None
        if height and weight and height > 0:
            bmi = round(weight / ((height / 100) ** 2), 1)

        pss_total = compute_pss(row, headers_idx)
        stress_index = round(pss_total / 40 * 100, 1) if pss_total is not None else None

        psqi_global = compute_psqi(row, headers_idx)
        sleep_index = round((1 - psqi_global / 21) * 100, 1) if psqi_global is not None else None

        nutrition_index, nutrition_answered = compute_nutrition(row, headers, headers_idx)

        wscore = wellness_score(stress_index, sleep_index, nutrition_index)

        # Skip junk rows: no age/height/weight and no scale scores at all
        # usually means the "Participant ID" cell just has stray text
        # (e.g. someone typed a note there) and the row has no real data.
        if age is None and height is None and weight is None and \
           pss_total is None and psqi_global is None and nutrition_index is None:
            continue

        results.append({
            "Participant ID": pid,
            "Weight": weight,
            "Height": height,
            "BMI": bmi,
            "Age": age,
            "PSS_total": pss_total,
            "PSS_category": pss_category(pss_total),
            "Stress_index": stress_index,
            "PSQI_global": psqi_global,
            "PSQI_category": psqi_category(psqi_global),
            "Sleep_index": sleep_index,
            "Nutrition_index": nutrition_index,
            "Nutrition_items_answered": nutrition_answered,
            "Wellness_score": wscore,
        })
    return results


# ---------------------------------------------------------------------
# GUI
# ---------------------------------------------------------------------
COLUMNS = [
    "Participant ID", "Weight", "Height", "BMI", "Age",
    "PSS_total", "PSS_category", "Stress_index",
    "PSQI_global", "PSQI_category", "Sleep_index",
    "Nutrition_index", "Nutrition_items_answered", "Wellness_score",
]


class App:
    def __init__(self, root):
        self.root = root
        self.root.title("HRVBP-HyperNet — Demographic Dashboard")
        self.root.geometry("1300x600")
        self.root.configure(bg="#f0f7ff")
        self.data = []

        # Configure style for Ttk components to match light blue/white theme
        style = ttk.Style(self.root)
        style.theme_use("clam")
        style.configure("Treeview", background="#ffffff", foreground="#1e293b", fieldbackground="#ffffff", rowheight=25)
        style.configure("Treeview.Heading", background="#e0f2fe", foreground="#1e40af", font=("Segoe UI", 10, "bold"))
        style.map("Treeview", background=[("selected", "#2563eb")], foreground=[("selected", "#ffffff")])

        top = tk.Frame(root, bg="#e0f2fe", height=90)
        top.pack(fill="x")
        top.pack_propagate(False)

        self.drop_label = tk.Label(
            top,
            text=("Drag & drop the questionnaire .xlsx file here"
                  if HAS_DND else
                  "Click 'Browse...' to load the questionnaire .xlsx file"),
            bg="#e0f2fe", fg="#1e40af", font=("Segoe UI", 13, "bold"),
        )
        self.drop_label.pack(side="left", padx=20, pady=10, expand=True)

        browse_btn = tk.Button(top, text="Browse...", command=self.browse_file,
                                font=("Segoe UI", 10, "bold"), bg="#2563eb", fg="white",
                                relief="flat", padx=12, pady=6)
        browse_btn.pack(side="right", padx=10)

        export_btn = tk.Button(top, text="Export CSV", command=self.export_csv,
                                font=("Segoe UI", 10, "bold"), bg="#0ea5e9", fg="white",
                                relief="flat", padx=12, pady=6)
        export_btn.pack(side="right", padx=10)

        self.status = tk.Label(root, text="No file loaded.", anchor="w",
                                font=("Segoe UI", 9), fg="#1e3a8a", bg="#f0f7ff")
        self.status.pack(fill="x", padx=10, pady=(4, 0))

        table_frame = tk.Frame(root, bg="#f0f7ff")
        table_frame.pack(fill="both", expand=True, padx=10, pady=10)

        self.tree = ttk.Treeview(table_frame, columns=COLUMNS, show="headings")
        for col in COLUMNS:
            self.tree.heading(col, text=col)
            width = 150 if col in ("Participant ID", "PSS_category", "PSQI_category") else 105
            self.tree.column(col, width=width, anchor="center")

        vsb = ttk.Scrollbar(table_frame, orient="vertical", command=self.tree.yview)
        hsb = ttk.Scrollbar(table_frame, orient="horizontal", command=self.tree.xview)
        self.tree.configure(yscrollcommand=vsb.set, xscrollcommand=hsb.set)

        self.tree.grid(row=0, column=0, sticky="nsew")
        vsb.grid(row=0, column=1, sticky="ns")
        hsb.grid(row=1, column=0, sticky="ew")
        table_frame.rowconfigure(0, weight=1)
        table_frame.columnconfigure(0, weight=1)

        if HAS_DND:
            self.root.drop_target_register(DND_FILES)
            self.root.dnd_bind("<<Drop>>", self.on_drop)

    def on_drop(self, event):
        path = event.data.strip("{}")
        self.load_file(path)

    def browse_file(self):
        path = filedialog.askopenfilename(filetypes=[("Excel files", "*.xlsx")])
        if path:
            self.load_file(path)

    def load_file(self, path):
        if not path.lower().endswith(".xlsx"):
            messagebox.showerror("Wrong file type", "Please provide an .xlsx file.")
            return
        try:
            self.status.config(text=f"Loading {os.path.basename(path)} ...")
            self.root.update_idletasks()
            self.data = process_workbook(path)
        except Exception as e:
            messagebox.showerror("Error loading file", str(e))
            self.status.config(text="Failed to load file.")
            return

        self.tree.delete(*self.tree.get_children())
        for row in self.data:
            values = [row[c] if row[c] is not None else "" for c in COLUMNS]
            self.tree.insert("", "end", values=values)

        self.status.config(
            text=f"Loaded {len(self.data)} participants from {os.path.basename(path)}"
        )

    def export_csv(self):
        if not self.data:
            messagebox.showwarning("No data", "Load a questionnaire file first.")
            return
        path = filedialog.asksaveasfilename(defaultextension=".csv",
                                             filetypes=[("CSV files", "*.csv")])
        if not path:
            return
        with open(path, "w", newline="", encoding="utf-8") as f:
            writer = csv.DictWriter(f, fieldnames=COLUMNS)
            writer.writeheader()
            writer.writerows(self.data)
        messagebox.showinfo("Exported", f"Saved to {path}")


def main():
    root = TkinterDnD.Tk() if HAS_DND else tk.Tk()
    App(root)
    root.mainloop()


if __name__ == "__main__":
    main()